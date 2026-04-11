"""
Import Only Nacharam Data for Metapharsic CRM
"""

import openpyxl
from openpyxl import Workbook
import re

def clean_contact(contact):
    """Clean phone number"""
    if not contact:
        return ""
    cleaned = re.sub(r'[^\d]', '', str(contact))
    if cleaned.startswith('91') and len(cleaned) > 10:
        cleaned = cleaned[2:]
    return cleaned

def process_nacharam_data():
    """Process all three files and filter ONLY Nacharam data"""
    print("🏥 Loading ONLY Nacharam Territory Data...")
    
    doctors = []
    pharmacies = []
    hospitals = []
    
    # 1. Process Healthcare Directory - filter Nacharam
    print("📊 Processing Healthcare Directory (Nacharam only)...")
    try:
        wb = openpyxl.load_workbook('Metapharsic_Healthcare_Directory_v3.xlsx')
        ws = wb['Healthcare Directory']
        
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0] or not str(row[0]).isdigit():
                continue
                
            area = str(row[3]) if row[3] else ""
            
            # ONLY Nacharam
            if area.lower() != 'nacharam':
                continue
                
            try:
                name = str(row[1]) if row[1] else ""
                entity_type = str(row[2]) if row[2] else ""
                address = str(row[4]) if row[4] else ""
                phone = clean_contact(row[5])
                rating = row[7]
                key_doctors = str(row[8]) if row[8] else ""
                speciality = str(row[9]) if row[9] else ""
                
                if not name:
                    continue
                
                # Calculate tier
                try:
                    rating_val = float(rating) if rating and rating != '—' else 0
                except:
                    rating_val = 0
                    
                tier = "A" if rating_val >= 4.5 else ("B" if rating_val >= 3.5 else "C")
                
                if entity_type.lower() in ['hospital', 'clinic']:
                    hospitals.append({
                        'name': name,
                        'type': entity_type,
                        'territory': 'Nacharam',
                        'tier': tier,
                        'phone': phone,
                        'address': address,
                        'speciality': speciality
                    })
                elif entity_type.lower() in ['dental']:
                    hospitals.append({
                        'name': name,
                        'type': 'Dental Clinic',
                        'territory': 'Nacharam',
                        'tier': tier,
                        'phone': phone,
                        'address': address,
                        'speciality': speciality
                    })
                elif entity_type.lower() in ['pharmacy']:
                    pharmacies.append({
                        'name': name,
                        'type': 'Independent Pharmacy' if 'Chain' not in name else 'Chain Pharmacy',
                        'territory': 'Nacharam',
                        'contact': phone,
                        'address': address,
                        'tier': tier
                    })
                elif 'RMP' in entity_type or 'GP' in entity_type:
                    doctors.append({
                        'name': name,
                        'specialty': speciality or 'General Practice',
                        'clinic': area,
                        'territory': 'Nacharam',
                        'tier': tier,
                        'contact': phone,
                        'address': address
                    })
                    
            except Exception as e:
                print(f"⚠️  Error: {e}")
                continue
    except Exception as e:
        print(f"❌ Error loading Healthcare Directory: {e}")
    
    print(f"✅ Healthcare Directory (Nacharam): {len(doctors)} doctors, {len(pharmacies)} pharmacies, {len(hospitals)} hospitals")
    
    # 2. Process Doctor List - ALL from this file are Nacharam
    print("👨‍️ Processing Doctor List...")
    try:
        wb = openpyxl.load_workbook('Nacharam_Hospital_DoctorList.xlsx')
        ws = wb['Nacharam Hospital Doctor List']
        
        current_hospital = ""
        
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]:
                continue
                
            cell_0 = str(row[0]) if row[0] else ""
            
            if '━━━━' in cell_0:
                current_hospital = cell_0.split('|')[0].replace('━━━━', '').strip()
                continue
            
            if not cell_0.isdigit():
                continue
                
            try:
                hospital = str(row[1]) if row[1] else current_hospital
                doctor_name = str(row[2]) if row[2] else ""
                qualification = str(row[3]) if row[3] else ""
                speciality = str(row[4]) if row[4] else ""
                mr_window = str(row[7]) if row[7] else ""
                
                if not doctor_name or 'Dr.' not in doctor_name:
                    continue
                
                doctors.append({
                    'name': doctor_name,
                    'specialty': speciality,
                    'clinic': hospital,
                    'territory': 'Nacharam',
                    'tier': 'A',
                    'contact': '',
                    'address': hospital,
                    'qualification': qualification,
                    'mr_visit_window': mr_window
                })
                
            except Exception as e:
                print(f"⚠️  Error: {e}")
                continue
    except Exception as e:
        print(f"❌ Error loading Doctor List: {e}")
    
    print(f"✅ Doctor List: {len(doctors)} doctors")
    
    # 3. Process Medical Halls - filter Nacharam
    print("🏪 Processing Medical Halls (Nacharam only)...")
    try:
        wb = openpyxl.load_workbook('Metapharsic_MedicalHalls_Directory.xlsx')
        ws = wb['Medical Halls & Pharmacies']
        
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0] or not str(row[0]).isdigit():
                continue
                
            area = str(row[9]) if row[9] else ""
            
            # ONLY Nacharam
            if area.lower() != 'nacharam':
                continue
                
            try:
                shop_name = str(row[1]) if row[1] else ""
                address = str(row[2]) if row[2] else ""
                contact = clean_contact(row[3])
                rating = row[5]
                shop_type = str(row[7]) if row[7] else ""
                notes = str(row[8]) if row[8] else ""
                
                if not shop_name:
                    continue
                
                # Calculate tier
                try:
                    rating_val = float(rating) if rating and rating != '—' else 0
                except:
                    rating_val = 0
                
                if '★' in notes or 'Priority' in notes:
                    tier = "A"
                elif shop_type.startswith("Chain"):
                    tier = "A"
                elif rating_val >= 4.5:
                    tier = "A"
                elif rating_val >= 3.5:
                    tier = "B"
                else:
                    tier = "C"
                
                pharmacies.append({
                    'name': shop_name,
                    'type': shop_type,
                    'territory': 'Nacharam',
                    'contact': contact,
                    'address': address,
                    'tier': tier,
                    'notes': notes
                })
                
            except Exception as e:
                print(f"⚠️  Error: {e}")
                continue
    except Exception as e:
        print(f"❌ Error loading Medical Halls: {e}")
    
    print(f"✅ Medical Halls (Nacharam): {len(pharmacies)} pharmacies")
    
    return doctors, pharmacies, hospitals

def create_upload_file(doctors, pharmacies, hospitals):
    """Create the upload-ready Excel file"""
    print("\n📦 Creating upload file...")
    
    wb = Workbook()
    
    # Pharmacies Sheet
    ws_pharmacies = wb.active
    ws_pharmacies.title = "Pharmacies"
    ws_pharmacies.append([
        "Name", "Owner", "Type", "Territory", "Contact", 
        "Email", "Address", "Tier"
    ])
    
    for ph in pharmacies:
        ws_pharmacies.append([
            ph['name'],
            ph.get('owner', ''),
            ph['type'],
            ph['territory'],
            ph.get('contact', ''),
            '',
            ph['address'],
            ph['tier']
        ])
    
    # Doctors Sheet
    ws_doctors = wb.create_sheet("Doctors")
    ws_doctors.append([
        "Name", "Specialty", "Clinic", "Territory", "Tier", 
        "Contact", "Email", "Address"
    ])
    
    for doc in doctors:
        ws_doctors.append([
            doc['name'],
            doc.get('specialty', ''),
            doc.get('clinic', ''),
            doc['territory'],
            doc['tier'],
            doc.get('contact', ''),
            '',
            doc.get('address', '')
        ])
    
    # Hospitals Sheet
    ws_hospitals = wb.create_sheet("Hospitals")
    ws_hospitals.append([
        "Name", "Type", "Territory", "Beds", "Contact", 
        "Email", "Address", "Tier"
    ])
    
    for hosp in hospitals:
        ws_hospitals.append([
            hosp['name'],
            hosp['type'],
            hosp['territory'],
            0,
            hosp.get('phone', ''),
            '',
            hosp['address'],
            hosp['tier']
        ])
    
    # MRs Sheet
    ws_mrs = wb.create_sheet("MRs")
    ws_mrs.append([
        "Name", "Territory", "Phone", "Email", "Base Salary", 
        "Daily Allowance", "Performance Score", "Total Sales", 
        "Targets Achieved", "Targets Missed"
    ])
    
    wb.save('Nacharam_Only_Upload.xlsx')
    print("✅ Saved to Nacharam_Only_Upload.xlsx")
    
    # Print summary
    print("\n" + "="*60)
    print("📊 NACHARAM DATA SUMMARY")
    print("="*60)
    print(f"🏪 Pharmacies: {len(pharmacies)}")
    print(f"👨‍️ Doctors: {len(doctors)}")
    print(f"🏥 Hospitals: {len(hospitals)}")
    print("="*60)

# Main execution
if __name__ == "__main__":
    print("="*60)
    print("🏥 LOADING NACHARAM TERRITORY DATA ONLY")
    print("="*60 + "\n")
    
    doctors, pharmacies, hospitals = process_nacharam_data()
    create_upload_file(doctors, pharmacies, hospitals)
    
    print("\n🚀 Upload Nacharam_Only_Upload.xlsx to Data Management!")
