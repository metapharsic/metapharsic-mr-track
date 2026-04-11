"""
Convert Metapharsic_MedicalHalls_Directory.xlsx to CRM-compatible format
This script transforms your pharmacy Excel into the correct column structure
"""

import openpyxl
from openpyxl import Workbook
import re

# Load your existing Excel file
input_file = "Metapharsic_MedicalHalls_Directory.xlsx"
output_file = "Metapharsic_Ready_To_Upload.xlsx"

print("📖 Loading your Excel file...")
wb_input = openpyxl.load_workbook(input_file)
ws_input = wb_input["Medical Halls & Pharmacies"]

# Create new workbook with multiple sheets
wb_output = Workbook()

# ============================================
# Sheet 1: Pharmacies (from your data)
# ============================================
ws_pharmacies = wb_output.active
ws_pharmacies.title = "Pharmacies"

# Headers for Pharmacies - MATCHING YOUR EXCEL STRUCTURE
pharmacy_headers = [
    "Name", "Owner", "Type", "Territory", "Contact", 
    "Email", "Address", "Tier"
]
ws_pharmacies.append(pharmacy_headers)

print("🏪 Processing Pharmacies...")

# Process each row (skip header rows and section headers)
row_count = 0
for row in ws_input.iter_rows(min_row=7, values_only=True):  # Start from row 7 (after headers)
    # Skip section headers like "🏥  HABSIGUDA"
    if not row[0] or str(row[0]).startswith("🏥") or not str(row[0]).isdigit():
        continue
    
    try:
        # Extract data from your Excel structure
        # Your columns: Sno | Name | Owner / Brand | Type | Territory | Contact | Email | Address | Tier
        sno = row[0] if row[0] else ""  # Skip section headers
        shop_name = str(row[1]) if row[1] else ""
        owner = str(row[2]) if row[2] else ""
        shop_type = str(row[3]) if row[3] else "Independent Medical Hall"
        territory = str(row[4]) if row[4] else ""
        contact = str(row[5]) if row[5] else ""
        email = str(row[6]) if row[6] else ""
        address = str(row[7]) if row[7] else ""
        tier = str(row[8]) if row[8] else "B"
        
        # Clean contact number (remove +91, spaces, dashes)
        contact_clean = re.sub(r'[^\d]', '', contact)
        if contact_clean.startswith('91') and len(contact_clean) > 10:
            contact_clean = contact_clean[2:]  # Remove country code
        
        # Clean tier (ensure it's A, B, or C)
        tier_clean = tier.strip().upper() if tier else "B"
        if tier_clean not in ['A', 'B', 'C']:
            tier_clean = "B"
        
        # Add to pharmacy sheet (matching database columns)
        ws_pharmacies.append([
            shop_name,           # Name
            owner,               # Owner
            shop_type,           # Type
            territory,           # Territory
            contact_clean,       # Contact
            email,               # Email
            address,             # Address
            tier_clean           # Tier
        ])
        row_count += 1
        
    except Exception as e:
        print(f"⚠️  Error processing row {row[0]}: {e}")
        continue

print(f"✅ Processed {row_count} pharmacies")

# ============================================
# Sheet 2: Doctors (empty template)
# ============================================
ws_doctors = wb_output.create_sheet("Doctors")
doctor_headers = [
    "Name", "Specialty", "Clinic", "Territory", "Tier", 
    "Contact", "Email", "Total Visits", "Total Orders", "Total Value"
]
ws_doctors.append(doctor_headers)
print("👨‍️ Created empty Doctors sheet (add your doctor data)")

# ============================================
# Sheet 3: Hospitals (empty template)
# ============================================
ws_hospitals = wb_output.create_sheet("Hospitals")
hospital_headers = [
    "Name", "Type", "City", "Beds", "Contact", 
    "Email", "Address", "Tier", "Total Purchases"
]
ws_hospitals.append(hospital_headers)
print("🏥 Created empty Hospitals sheet (add your hospital data)")

# ============================================
# Sheet 4: MRs (empty template)
# ============================================
ws_mrs = wb_output.create_sheet("MRs")
mr_headers = [
    "Name", "Territory", "Phone", "Email", "Base Salary", 
    "Daily Allowance", "Performance Score", "Total Sales", 
    "Targets Achieved", "Targets Missed"
]
ws_mrs.append(mr_headers)
print("👔 Created empty MRs sheet (add your MR data)")

# ============================================
# Save the output file
# ============================================
wb_output.save(output_file)

print("\n" + "="*60)
print("✅ CONVERSION COMPLETE!")
print("="*60)
print(f"\n📁 Output file: {output_file}")
print(f"📊 Sheets created:")
print(f"   1. Pharmacies ({row_count} records from your data)")
print(f"   2. Doctors (empty - add your doctor data)")
print(f"   3. Hospitals (empty - add your hospital data)")
print(f"   4. MRs (empty - add your MR data)")
print("\n🚀 Next Steps:")
print("   1. Open the output file")
print("   2. Fill in Doctors, Hospitals, and MRs sheets if needed")
print("   3. Go to http://localhost:3000/data-management")
print("   4. Upload the file!")
print("="*60)
