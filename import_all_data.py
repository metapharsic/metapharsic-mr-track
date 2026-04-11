"""
Smart Excel Import Script for Metapharsic CRM
Converts multiple Excel files to upload-ready format
"""

import openpyxl
from openpyxl import Workbook
import re

def clean_contact(contact):
    """Clean phone number"""
    if not contact:
        return ""
    cleaned = re.sub(r'[^\d]', '', str(contact))
    if cleaned.startswith('91') and len(cleaned) > 10:
        cleaned = cleaned[2:]
    return cleaned

def clean_tier(tier):
    """Ensure tier is A, B, or C"""
    if not tier:
        return "B"
    tier_str = str(tier).strip().upper()
    if tier_str in ['A', 'B', 'C']:
        return tier_str
    return "B"

def process_healthcare_directory():
    """Process Metapharsic_Healthcare_Directory_v3.xlsx"""
    print("📊 Processing Healthcare Directory...")
    
    wb = openpyxl.load_workbook('Metapharsic_Healthcare_Directory_v3.xlsx')
    ws = wb['Healthcare Directory']
    
    doctors = []
    pharmacies = []
    hospitals = []
    
    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip header rows
        if not row[0] or not str(row[0]).isdigit():
            continue
            
        try:
            name = str(row[1]) if row[1] else ""
            entity_type = str(row[2]) if row[2] else ""
            area = str(row[3]) if row[3] else ""
            address = str(row[4]) if row[4] else ""
            phone = clean_contact(row[5])
            timings = str(row[6]) if row[6] else ""
            rating = row[7]
            key_doctors = str(row[8]) if row[8] else ""
            speciality = str(row[9]) if row[9] else ""
            
            if not name:
                continue
            
            # Calculate tier based on rating
            try:
                rating_val = float(rating) if rating and rating != '—' else 0
            except:
                rating_val = 0
                
            tier = "A" if rating_val >= 4.5 else ("B" if rating_val >= 3.5 else "C")
            
            # Classify by entity type
            if entity_type.lower() in ['hospital', 'clinic']:
                hospitals.append({
                    'name': name,
                    'type': entity_type,
                    'territory': area,
                    'tier': tier,
                    'phone': phone,
                    'address': address,
                    'timings': timings,
                    'key_doctors': key_doctors,
                    'speciality': speciality,
                    'rating': rating_val
                })
            elif entity_type.lower() in ['dental']:
                # Dental clinics go to hospitals table with type 'Dental'
                hospitals.append({
                    'name': name,
                    'type': 'Dental Clinic',
                    'territory': area,
                    'tier': tier,
                    'phone': phone,
                    'address': address,
                    'timings': timings,
                    'key_doctors': key_doctors,
                    'speciality': speciality,
                    'rating': rating_val
                })
            elif entity_type.lower() in ['pharmacy']:
                pharmacies.append({
                    'name': name,
                    'owner': '',
                    'type': 'Independent Pharmacy' if 'Chain' not in name else 'Chain Pharmacy',
                    'territory': area,
                    'contact': phone,
                    'email': '',
                    'address': address,
                    'tier': tier,
                    'rating': rating_val
                })
            elif 'RMP' in entity_type or 'GP' in entity_type:
                # RMP/GP goes to doctors table
                doctors.append({
                    'name': name,
                    'specialty': speciality or 'General Practice',
                    'clinic': area,
                    'territory': area,
                    'tier': tier,
                    'contact': phone,
                    'email': '',
                    'address': address,
                    'isRMP': True
                })
            else:
                # Check if it has doctors listed
                if key_doctors and 'Dr.' in key_doctors:
                    # It's a hospital/clinic with doctors
                    hospitals.append({
                        'name': name,
                        'type': entity_type,
                        'territory': area,
                        'tier': tier,
                        'phone': phone,
                        'address': address,
                        'timings': timings,
                        'key_doctors': key_doctors,
                        'speciality': speciality,
                        'rating': rating_val
                    })
                    
        except Exception as e:
            print(f"⚠️  Error processing row: {e}")
            continue
    
    print(f"✅ Healthcare Directory: {len(doctors)} doctors, {len(pharmacies)} pharmacies, {len(hospitals)} hospitals")
    return doctors, pharmacies, hospitals

def process_doctor_list():
    """Process Nacharam_Hospital_DoctorList.xlsx"""
    print("👨‍️ Processing Doctor List...")
    
    wb = openpyxl.load_workbook('Nacharam_Hospital_DoctorList.xlsx')
    ws = wb['Nacharam Hospital Doctor List']
    
    doctors = []
    current_hospital = ""
    
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
            
        cell_0 = str(row[0]) if row[0] else ""
        
        # Detect hospital header
        if '━━━━' in cell_0:
            # Extract hospital name
            current_hospital = cell_0.split('|')[0].replace('━━━━', '').strip()
            continue
        
        # Skip if not a number
        if not cell_0.isdigit():
            continue
            
        try:
            hospital = str(row[1]) if row[1] else current_hospital
            doctor_name = str(row[2]) if row[2] else ""
            qualification = str(row[3]) if row[3] else ""
            speciality = str(row[4]) if row[4] else ""
            dept = str(row[5]) if row[5] else ""
            opd_timing = str(row[6]) if row[6] else ""
            mr_window = str(row[7]) if row[7] else ""
            notes = str(row[8]) if row[8] else ""
            
            if not doctor_name or 'Dr.' not in doctor_name:
                continue
            
            doctors.append({
                'name': doctor_name,
                'specialty': speciality or dept,
                'clinic': hospital,
                'territory': 'Nacharam',  # All doctors from this file are in Nacharam
                'tier': 'A',  # Hospital doctors are typically A-tier
                'contact': '',
                'email': '',
                'address': hospital,
                'qualification': qualification,
                'opd_timing': opd_timing,
                'mr_visit_window': mr_window,
                'notes': notes,
                'isRMP': False
            })
            
        except Exception as e:
            print(f"⚠️  Error processing doctor row: {e}")
            continue
    
    print(f"✅ Doctor List: {len(doctors)} doctors")
    return doctors

def process_medical_halls():
    """Process Metapharsic_MedicalHalls_Directory.xlsx"""
    print("🏪 Processing Medical Halls...")
    
    wb = openpyxl.load_workbook('Metapharsic_MedicalHalls_Directory.xlsx')
    ws = wb['Medical Halls & Pharmacies']
    
    pharmacies = []
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0] or not str(row[0]).isdigit():
            continue
            
        try:
            shop_name = str(row[1]) if row[1] else ""
            address = str(row[2]) if row[2] else ""
            contact = clean_contact(row[3])
            shop_hours = str(row[4]) if row[4] else ""
            rating = row[5]
            mr_window = str(row[6]) if row[6] else ""
            shop_type = str(row[7]) if row[7] else ""
            notes = str(row[8]) if row[8] else ""
            area = str(row[9]) if row[9] else ""
            
            if not shop_name:
                continue
            
            # Calculate tier
            try:
                rating_val = float(rating) if rating and rating != '—' else 0
            except:
                rating_val = 0
            
            # Priority shops marked with ★ get A tier
            if '★' in notes or 'Priority' in notes:
                tier = "A"
            elif shop_type.startswith("Chain"):
                tier = "A"
            elif rating_val >= 4.5:
                tier = "A"
            elif rating_val >= 3.5:
                tier = "B"
            else:
                tier = "C"
            
            pharmacies.append({
                'name': shop_name,
                'owner': '',
                'type': shop_type,
                'territory': area,
                'contact': contact,
                'email': '',
                'address': address,
                'tier': tier,
                'mr_visit_window': mr_window,
                'notes': notes,
                'rating': rating_val
            })
            
        except Exception as e:
            print(f"⚠️  Error processing pharmacy row: {e}")
            continue
    
    print(f"✅ Medical Halls: {len(pharmacies)} pharmacies")
    return pharmacies

def create_upload_file(all_doctors, all_pharmacies, all_hospitals):
    """Create the upload-ready Excel file"""
    print("\n📦 Creating upload file...")
    
    wb = Workbook()
    
    # Pharmacies Sheet
    ws_pharmacies = wb.active
    ws_pharmacies.title = "Pharmacies"
    ws_pharmacies.append([
        "Name", "Owner", "Type", "Territory", "Contact", 
        "Email", "Address", "Tier"
    ])
    
    for ph in all_pharmacies:
        ws_pharmacies.append([
            ph['name'],
            ph.get('owner', ''),
            ph['type'],
            ph['territory'],
            ph.get('contact', ''),
            ph.get('email', ''),
            ph['address'],
            ph['tier']
        ])
    
    # Doctors Sheet
    ws_doctors = wb.create_sheet("Doctors")
    ws_doctors.append([
        "Name", "Specialty", "Clinic", "Territory", "Tier", 
        "Contact", "Email", "Address"
    ])
    
    for doc in all_doctors:
        ws_doctors.append([
            doc['name'],
            doc.get('specialty', ''),
            doc.get('clinic', ''),
            doc['territory'],
            doc['tier'],
            doc.get('contact', ''),
            doc.get('email', ''),
            doc.get('address', '')
        ])
    
    # Hospitals Sheet
    ws_hospitals = wb.create_sheet("Hospitals")
    ws_hospitals.append([
        "Name", "Type", "Territory", "Beds", "Contact", 
        "Email", "Address", "Tier"
    ])
    
    for hosp in all_hospitals:
        ws_hospitals.append([
            hosp['name'],
            hosp['type'],
            hosp['territory'],
            0,  # Beds - not available
            hosp.get('phone', ''),
            '',
            hosp['address'],
            hosp['tier']
        ])
    
    # MRs Sheet (empty for now)
    ws_mrs = wb.create_sheet("MRs")
    ws_mrs.append([
        "Name", "Territory", "Phone", "Email", "Base Salary", 
        "Daily Allowance", "Performance Score", "Total Sales", 
        "Targets Achieved", "Targets Missed"
    ])
    
    wb.save('Metapharsic_Upload_Ready.xlsx')
    print("✅ Saved to Metapharsic_Upload_Ready.xlsx")
    
    # Print summary
    print("\n" + "="*60)
    print("📊 SUMMARY")
    print("="*60)
    print(f"Pharmacies: {len(all_pharmacies)}")
    
    # Count by territory
    territories = {}
    for ph in all_pharmacies:
        t = ph['territory']
        territories[t] = territories.get(t, 0) + 1
    for t, count in sorted(territories.items()):
        print(f"  - {t}: {count}")
    
    print(f"\nDoctors: {len(all_doctors)}")
    doc_territories = {}
    for doc in all_doctors:
        t = doc['territory']
        doc_territories[t] = doc_territories.get(t, 0) + 1
    for t, count in sorted(doc_territories.items()):
        print(f"  - {t}: {count}")
    
    print(f"\nHospitals: {len(all_hospitals)}")
    hosp_territories = {}
    for hosp in all_hospitals:
        t = hosp['territory']
        hosp_territories[t] = hosp_territories.get(t, 0) + 1
    for t, count in sorted(hosp_territories.items()):
        print(f"  - {t}: {count}")
    
    print("\n🚀 Next: Upload Metapharsic_Upload_Ready.xlsx to Data Management!")
    print("="*60)

# Main execution
if __name__ == "__main__":
    print("="*60)
    print("🏥 METAPHARSIC HEALTHCARE DATA IMPORTER")
    print("="*60 + "\n")
    
    # Process all three files
    hc_doctors, hc_pharmacies, hc_hospitals = process_healthcare_directory()
    doc_list_doctors = process_doctor_list()
    medical_halls_pharmacies = process_medical_halls()
    
    # Combine all data
    all_doctors = hc_doctors + doc_list_doctors
    all_pharmacies = hc_pharmacies + medical_halls_pharmacies
    all_hospitals = hc_hospitals
    
    # Create upload file
    create_upload_file(all_doctors, all_pharmacies, all_hospitals)
